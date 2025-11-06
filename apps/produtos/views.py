# apps/produtos/views.py
import io
from pyexpat.errors import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView,
    TemplateView, FormView, View
)
from datetime import date, timedelta
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseBadRequest, JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from django.db.models import Q
from datetime import date, datetime, timedelta
import json
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl import Workbook
import pandas as pd
from apps.analytics import models
from apps.core.views import BaseMPAView
from apps.servicos.models import Servico
from .models import (
    Categoria, Fabricante,
    Produto, Lote, HistoricoPreco
)
from django.db.models import Q
from .forms import ImportarProdutosForm, LoteForm, ProdutoForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import redirect
from apps.core.models import Categoria
from .forms import CategoriaForm
from django.http import JsonResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import models
from .models import Produto
from apps.core.models import Empresa
from django.db import models
from django.http import JsonResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from apps.produtos.models import Produto
from apps.core.models import Empresa
from django.contrib.auth.mixins import AccessMixin
import logging
import pandas as pd
from decimal import Decimal, InvalidOperation
from django.utils import timezone
from django.core.exceptions import ValidationError
from apps.fornecedores.models import Fornecedor
import openpyxl
from openpyxl import Workbook
from apps.core.models import Categoria




class PermissaoAcaoMixin(AccessMixin):
    # CRÍTICO: Definir esta variável na View
    acao_requerida = None 

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()

        try:
            # Tenta obter o Funcionario (ligação fundamental)
            funcionario = request.user.funcionario 
        except Exception:
            messages.error(request, "Acesso negado. O seu usuário não está ligado a um registro de funcionário.")
            return self.handle_no_permission()

        if self.acao_requerida:
            # Usa a lógica dinâmica do modelo Funcionario (que já criámos)
            if not funcionario.pode_realizar_acao(self.acao_requerida):
                messages.error(request, f"Acesso negado. O seu cargo não permite realizar a ação de '{self.acao_requerida}'.")
                return redirect(reverse_lazy('core:dashboard'))

        return super().dispatch(request, *args, **kwargs)




@login_required
def dashboard_view(request):
    """View do dashboard com notificações de estoque e validade"""
    hoje = date.today()
    data_limite = hoje + timedelta(days=30)
    
    # Lotes vencendo
    lotes_vencendo = Lote.objects.filter(
        validade__gte=hoje,
        validade__lte=data_limite
    ).select_related('produto').order_by('validade')
    
    # Produtos com estoque baixo
    produtos_ativos = Produto.objects.filter(ativo=True)
    produtos_estoque_baixo = [p for p in produtos_ativos if p.estoque_baixo]
    
    # Construir notificações
    lista_notificacoes = []
    
    for lote in lotes_vencendo:
        lista_notificacoes.append({
            'mensagem': f"Lote {lote.numero_lote} de {lote.produto.nome_produto} vence em {lote.validade.strftime('%d/%m/%Y')}",
            'detalhe': f"{lote.quantidade} unidades restantes",
            'css_class': 'bg-yellow-400'
        })

    for produto in produtos_estoque_baixo:
        lista_notificacoes.append({
            'mensagem': f"Estoque baixo: {produto.nome_produto}",
            'detalhe': f"Apenas {produto.estoque_atual} unidades",
            'css_class': 'bg-red-500'
        })
    
    context = {
        'total_notificacoes': len(lista_notificacoes),
        'lista_notificacoes': lista_notificacoes,
    }

    return render(request, 'core/dashboard.html', context)


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from .models import Produto
import json




@login_required
@require_http_methods(["GET"])
def buscar_produtos_api(request):
    """
    API para buscar produtos no PDV - agora suporta listar todos
    """
    try:
        busca = request.GET.get('q', '').strip()
        categoria_id = request.GET.get('categoria', '').strip()
        
        # Buscar produtos da empresa do usuário
        if hasattr(request.user, 'funcionario') and request.user.funcionario:
            empresa = request.user.funcionario.empresa
        else:
            return JsonResponse({
                'success': False,
                'message': 'Empresa não encontrada'
            }, status=400)
        
        # Query base
        produtos = Produto.objects.filter(
            empresa=empresa,
            ativo=True
        ).select_related('categoria', 'fornecedor', 'fabricante')
        
        # Filtrar por categoria se especificado
        if categoria_id and categoria_id != 'todos':
            produtos = produtos.filter(categoria_id=categoria_id)
        
        # Filtrar por busca se especificado
        if busca:
            produtos = produtos.filter(
                Q(codigo_barras__icontains=busca) |
                Q(codigo_interno__icontains=busca) |
                Q(nome_produto__icontains=busca)
            )
        
        # Limitar resultados apenas se houver busca específica
        if busca:
            produtos = produtos[:20]
        else:
            produtos = produtos[:100]  # Limite maior para listagem geral
        
        produtos_data = []
        for produto in produtos:
            produtos_data.append({
                'id': produto.id,
                'codigo_interno': produto.codigo_interno,
                'codigo_barras': produto.codigo_barras,
                'nome_produto': produto.nome_produto,
                'categoria': produto.categoria.nome if produto.categoria else '',
                'categoria_id': produto.categoria.id if produto.categoria else None,
                'fornecedor': str(produto.fornecedor) if produto.fornecedor else '',
                'fabricante': str(produto.fabricante) if produto.fabricante else '',
                'estoque_atual': produto.estoque_atual,
                'estoque_minimo': produto.estoque_minimo,
                'preco_custo': float(produto.preco_custo),
                'preco_venda': float(produto.preco_venda),
                'margem_lucro': float(produto.margem_lucro),
                'foto_url': produto.foto.url if produto.foto else None,
                'valor_estoque': float(produto.valor_estoque),
                'estoque_baixo': produto.estoque_atual <= produto.estoque_minimo,
                'disponivel': produto.estoque_atual > 0,
                'iva_percentual': getattr(produto, 'iva_percentual', 0)
            })
        
        return JsonResponse({
            'success': True,
            'produtos': produtos_data,
            'total_encontrados': len(produtos_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }, status=500)



class ProdutosView(LoginRequiredMixin, ListView):
    template_name = 'produtos/produto_list.html'
    model = Produto
    context_object_name = 'produtos'
    paginate_by = 20

    def get_empresa(self):
        """ Método seguro para obter a empresa do utilizador logado. """
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario and user.funcionario.empresa:
            return user.funcionario.empresa
        if user.is_superuser:
            return Empresa.objects.first()
        return None

    def get_queryset(self):
        """
        Filtra e ordena a lista de produtos.
        """
        empresa = self.get_empresa()
        if not empresa:
            return Produto.objects.none()

        # Começa com o queryset base
        queryset = Produto.objects.filter(empresa=empresa).select_related('categoria')
        
        # Filtros do URL
        search = self.request.GET.get('search', '')
        categoria_filter = self.request.GET.get('categoria', '')
        status_filter = self.request.GET.get('status', '')

        if search:
            queryset = queryset.filter(
                Q(nome_produto__icontains=search) |
                Q(codigo_barras__icontains=search)
            )
        
        if categoria_filter:
            queryset = queryset.filter(categoria_id=categoria_filter)
        
        if status_filter == 'ativo':
            queryset = queryset.filter(ativo=True)
        elif status_filter == 'inativo':
            queryset = queryset.filter(ativo=False)
        elif status_filter == 'estoque_baixo':
            queryset = queryset.filter(estoque_atual__lte=F('estoque_minimo'))
        
        return queryset.order_by('nome_produto')

    def get_context_data(self, **kwargs):
        """ Adiciona dados extra (estatísticas, etc.) ao contexto. """
        context = super().get_context_data(**kwargs)
        empresa = self.get_empresa()

        if empresa:
            todas_categorias = Categoria.objects.filter(empresa=empresa).order_by('nome')
            servicos = Servico.objects.filter(empresa=empresa)
            todos_produtos = Produto.objects.filter(empresa=empresa)
            
            context['produtos_stats'] = {
                'total': todos_produtos.count(),
                'ativos': todos_produtos.filter(ativo=True).count(),
                'categorias': todas_categorias.count(),
                'servicos': servicos.count(),
            }
            context['categorias'] = todas_categorias.filter(ativa=True)
        
        return context




class CriarProdutoView(LoginRequiredMixin, CreateView):
    model = Produto
    form_class = ProdutoForm
    template_name = 'produtos/produto_form.html'
    success_url = reverse_lazy('produtos:produto_list')
    acao_requerida = 'editar_produtos'

    def get_empresa(self):
        """ Método seguro para obter a empresa do utilizador logado. """
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario and user.funcionario.empresa:
            return user.funcionario.empresa
        if user.is_superuser:
            from apps.core.models import Empresa
            return Empresa.objects.first()
        return None

    def dispatch(self, request, *args, **kwargs):
        """ Verifica se o utilizador tem uma empresa antes de continuar. """
        if not self.get_empresa():
            messages.error(request, "O seu utilizador não está associado a nenhuma empresa.")
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        """ Passa a empresa para o formulário para filtrar os ForeignKeys. """
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa()
        return kwargs

    def form_valid(self, form):
        produto = form.save(commit=False)
        produto.empresa = self.get_empresa()
        produto.save()
        messages.success(self.request, f"Produto '{produto.nome_produto}' criado com sucesso!")
        
        # MELHORIA: Deixe a CreateView tratar do redirecionamento
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Adicionar Novo Produto"
        return context


class EditarProdutoView(LoginRequiredMixin, UpdateView):
    model = Produto
    form_class = ProdutoForm
    template_name = 'produtos/produto_form.html'
    success_url = reverse_lazy('produtos:produto_list') # Redireciona para a lista de produtos
    pk_url_kwarg = 'produto_id' # Informa que o ID na URL é 'produto_id'
    acao_requerida = 'editar_produtos'

    def get_queryset(self):
        """ Garante que o utilizador só pode editar produtos da sua própria empresa. """
        empresa = self.request.user.funcionario.empresa
        return Produto.objects.filter(empresa=empresa)

    def get_form_kwargs(self):
        """ Passa a empresa para o formulário para filtrar os campos ForeignKey. """
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.request.user.funcionario.empresa
        return kwargs

    def form_valid(self, form):
        """ Adiciona uma mensagem de sucesso antes de redirecionar. """
        messages.success(self.request, f"Produto '{form.instance.nome_produto}' atualizado com sucesso!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Editar Produto: {self.object.nome_produto}"
        return context
    

class DeletarProdutoView(LoginRequiredMixin, View):
    acao_requerida = 'editar_produtos'
    def post(self, request, produto_id):
        try:
            empresa = self.get_empresa(request)
            from apps.produtos.models import Produto
            produto = Produto.objects.get(id=produto_id, empresa=empresa)
            
            nome = produto.nome_produto
            produto.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Produto "{nome}" removido com sucesso'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    def get_empresa(self, request):
        if hasattr(request.user, 'usuario') and request.user.usuario.empresa:
            return request.user.usuario.empresa
        elif hasattr(request.user, 'profile') and request.user.profile.empresa:
            return request.user.profile.empresa
        else:
            from apps.core.models import Empresa
            return Empresa.objects.first()


class ToggleProdutoView(LoginRequiredMixin, View):
    acao_requerida = 'editar_produtos'
    def post(self, request, produto_id):
        try:
            empresa = self.get_empresa(request)
            from apps.produtos.models import Produto
            produto = Produto.objects.get(id=produto_id, empresa=empresa)
            
            produto.ativo = not produto.ativo
            produto.save()
            
            return JsonResponse({
                'success': True,
                'ativo': produto.ativo,
                'message': f'Produto {"ativado" if produto.ativo else "desativado"} com sucesso'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    def get_empresa(self, request):
        if hasattr(request.user, 'usuario') and request.user.usuario.empresa:
            return request.user.usuario.empresa
        elif hasattr(request.user, 'profile') and request.user.profile.empresa:
            return request.user.profile.empresa
        else:
            from apps.core.models import Empresa
            return Empresa.objects.first()


# apps/produtos/views.py (substituir a TemplateProdutosView existente)

class TemplateProdutosView(LoginRequiredMixin, View):
    def get(self, request):
        try:
            empresa = self.get_empresa()
            if not empresa:
                return HttpResponseBadRequest('Empresa não encontrada')
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos"
            
            # Definir cabeçalhos
            headers = [
                'nome_produto', 'codigo_barras', 'categoria', 'preco_custo', 'preco_venda',
                'nome_comercial', 'codigo_interno', 'fabricante', 'fornecedor',
                'estoque_atual', 'estoque_minimo', 'estoque_maximo', 'margem_lucro',
                'desconto_percentual', 'observacoes', 'ativo'
            ]
            
            # Adicionar cabeçalhos
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Adicionar dados de exemplo
            exemplo = [
                'Paracetamol 500mg',    # nome_produto
                '7891234567890',        # codigo_barras
                'Medicamentos',         # categoria
                '2.50',                 # preco_custo
                '5.00',                 # preco_venda
                'Paracetamol 500mg',    # nome_comercial
                'PARA500',              # codigo_interno
                'EMS',                  # fabricante
                'Distribuidora ABC',    # fornecedor
                '100',                  # estoque_atual
                '10',                   # estoque_minimo
                '1000',                 # estoque_maximo
                '100.00',               # margem_lucro
                '0.00',                 # desconto_percentual
                'Produto de exemplo',   # observacoes
                '1'                     # ativo
            ]
            
            for col, value in enumerate(exemplo, 1):
                ws.cell(row=2, column=col, value=value)
            
            # Adicionar instruções em uma nova planilha
            ws_instrucoes = wb.create_sheet("Instruções")
            
            instrucoes = [
                "INSTRUÇÕES PARA IMPORTAÇÃO DE PRODUTOS",
                "",
                "CAMPOS OBRIGATÓRIOS:",
                "- nome_produto: Nome do produto",
                "- codigo_barras: Código de barras único",
                "- categoria: Nome da categoria",
                "- preco_custo: Preço de custo (formato: 12.50)",
                "- preco_venda: Preço de venda (formato: 25.00)",
                "",
                "CAMPOS OPCIONAIS:",
                "- nome_comercial: Nome comercial (se vazio, usará nome_produto)",
                "- codigo_interno: Código interno (se vazio, usará codigo_barras)",
                "- fabricante: Nome do fabricante",
                "- fornecedor: Nome do fornecedor (deve existir no sistema)",
                "- estoque_atual: Quantidade em estoque (padrão: 0)",
                "- estoque_minimo: Estoque mínimo (padrão: 1)",
                "- estoque_maximo: Estoque máximo (padrão: 100)",
                "- margem_lucro: Margem de lucro em % (padrão: 0)",
                "- desconto_percentual: Desconto em % (padrão: 0)",
                "- observacoes: Observações sobre o produto",
                "- ativo: 1 para ativo, 0 para inativo (padrão: 1)",
                "",
                "FORMATOS:",
                "- Preços: Use ponto como separador decimal (ex: 12.50)",
                "- Booleanos: Use 1/0 ou true/false ou sim/não",
                "- Texto: Evite caracteres especiais",
                "",
                "NOTAS IMPORTANTES:",
                "- Códigos de barras devem ser únicos",
                "- Categorias e fabricantes serão criados automaticamente se não existirem",
                "- Fornecedores devem existir previamente no sistema",
                "- Linhas com erros serão ignoradas",
                "- Use a opção 'Validar apenas' para testar antes de importar"
            ]
            
            for row, instrucao in enumerate(instrucoes, 1):
                ws_instrucoes.cell(row=row, column=1, value=instrucao)
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Salvar em buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Retornar response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="template_produtos_{empresa.nome}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            
            return response
            
        except Exception as e:
            logger.error(f"Erro ao gerar template: {str(e)}")
            return HttpResponseBadRequest(f'Erro ao gerar template: {str(e)}')

    def get_empresa(self):
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario and user.funcionario.empresa:
            return user.funcionario.empresa
        if user.is_superuser:
            return Empresa.objects.first()
        return None


class CategoriaBaseView(LoginRequiredMixin):
    """
    View base para garantir que as operações são feitas na empresa correta.
    Agora com métodos seguros para obter a empresa.
    """
    model = Categoria

    def get_empresa(self):
        """ Método seguro para obter a empresa do utilizador logado. """
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario and user.funcionario.empresa:
            return user.funcionario.empresa
        # Adicione um fallback para o superuser ou outros casos, se necessário
        if user.is_superuser and not hasattr(user, 'funcionario'):
             from apps.core.models import Empresa
             return Empresa.objects.first()
        return None

    def dispatch(self, request, *args, **kwargs):
        """ Garante que o utilizador tem uma empresa antes de prosseguir. """
        if not self.get_empresa():
            messages.error(request, "O seu utilizador não está associado a nenhuma empresa.")
            return redirect('core:dashboard') # Redireciona para o dashboard principal
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        """ Filtra as categorias para mostrar apenas as da empresa do utilizador. """
        return Categoria.objects.filter(empresa=self.get_empresa())

class CategoriaListView(CategoriaBaseView, ListView):
    template_name = 'produtos/categoria_list.html'
    context_object_name = 'categorias'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Categorias de Produtos"
        return context



class CategoriaCreateView(CategoriaBaseView, CreateView):
    model = Categoria
    # --- PONTO CRÍTICO DA CORREÇÃO ---
    form_class = CategoriaForm
    
    template_name = 'produtos/categoria_form.html'
    success_url = reverse_lazy('produtos:categoria_list')

    def get_form_kwargs(self):
        # Este método é essencial para passar a 'empresa' para o formulário
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa()
        return kwargs

    def form_valid(self, form):
        # A lógica aqui está correta, associa a empresa antes de salvar
        form.instance.empresa = self.get_empresa()
        messages.success(self.request, "Categoria criada com sucesso!")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Adicionar Nova Categoria"
        return context       



class CategoriaUpdateView(CategoriaBaseView, UpdateView):
    form_class = CategoriaForm
    template_name = 'produtos/categoria_form.html'
    success_url = reverse_lazy('produtos:categoria_list')

    def get_form_kwargs(self):
        """ Passa a empresa para o formulário. """
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa()
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, "Categoria atualizada com sucesso!")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Editar Categoria: {self.object.nome}"
        return context

class CategoriaDeleteView(CategoriaBaseView, DeleteView):
    template_name = 'produtos/categoria_confirm_delete.html'
    success_url = reverse_lazy('produtos:categoria_lista')
    
    def post(self, request, *args, **kwargs):
        # O método post é o correto para adicionar a mensagem antes de apagar
        self.object = self.get_object()
        messages.success(self.request, f"Categoria '{self.object.nome}' eliminada com sucesso!")
        return self.delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Eliminar Categoria: {self.object.nome}"
        return context   



#######################################
class FabricanteListView(LoginRequiredMixin, ListView):
    model = Fabricante
    template_name = "produtos/fabricante_list.html"
    context_object_name = "fabricantes"
    ordering = ['nome']

class FabricanteDetailView(LoginRequiredMixin, DetailView):
    model = Fabricante
    template_name = "produtos/fabricante_detail.html"

class FabricanteCreateView(LoginRequiredMixin, CreateView):
    model = Fabricante
    fields = '__all__'
    template_name = "produtos/form.html"
    success_url = reverse_lazy("produtos:fabricante_list")

class FabricanteUpdateView(LoginRequiredMixin, UpdateView):
    model = Fabricante
    fields = '__all__'
    template_name = "produtos/form.html"
    success_url = reverse_lazy("produtos:fabricante_list")

class FabricanteDeleteView(LoginRequiredMixin, DeleteView):
    model = Fabricante
    template_name = "produtos/confirm_delete.html"
    success_url = reverse_lazy("produtos:fabricante_list")


# ===============================
# LOTE VIEWS
# ===============================

class LoteListView(LoginRequiredMixin, ListView):
    model = Lote
    template_name = "produtos/lote_list.html"
    context_object_name = "lotes"
    paginate_by = 20  # Define a paginação aqui

    def get_queryset(self):
        """
        Este método agora constrói a consulta ao banco de dados de forma dinâmica,
        aplicando os filtros e a ordenação recebidos do formulário.
        """
        # Começa com todos os lotes (respeitando permissões, se aplicável)
        queryset = Lote.objects.select_related('produto').all()

        # Obtém os parâmetros do URL (ex: ?search=paracetamol&status=vencendo)
        search_query = self.request.GET.get('search', '').strip()
        status_filter = self.request.GET.get('status', '')
        sort_by = self.request.GET.get('sort', '-data_validade') # Padrão: validade mais longe

        # 1. Aplicar filtro de pesquisa
        if search_query:
            queryset = queryset.filter(
                Q(produto__nome_produto__icontains=search_query) |
                Q(numero_lote__icontains=search_query)
            )

        # 2. Aplicar filtro de status de validade
        hoje = date.today()
        if status_filter == 'vencido':
            queryset = queryset.filter(data_validade__lt=hoje)
        elif status_filter == 'vencendo':
            data_limite = hoje + timedelta(days=30)
            queryset = queryset.filter(data_validade__gte=hoje, data_validade__lte=data_limite)

        # 3. Aplicar ordenação
        # Validar as opções de ordenação para segurança
        valid_sort_options = ['data_validade', '-data_validade', 'quantidade_atual', '-quantidade_atual']
        if sort_by in valid_sort_options:
            queryset = queryset.order_by(sort_by)

        return queryset

    def get_context_data(self, **kwargs):
        """ Adiciona os valores dos filtros ao contexto para manter o estado no template. """
        context = super().get_context_data(**kwargs)
        # Passa os valores atuais dos filtros de volta para o template
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_sort'] = self.request.GET.get('sort', '-data_validade')
        context['add_product_form'] = ProdutoForm() 
        
        
        return context


class LoteDetailView(LoginRequiredMixin, DetailView):
    model = Lote
    template_name = "produtos/lote_detail.html"






from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView

from .models import Lote
from .forms import LoteForm, ProdutoForm # Garanta que os forms estão importados

# --- BASE VIEW PARA LOTES (BOA PRÁTICA) ---
class LoteBaseView(LoginRequiredMixin):
    """ View base para Lotes, garante o modelo e o filtro por empresa. """
    model = Lote

    def get_queryset(self):
        # Garante que qualquer view que herde disto só aceda a lotes da empresa do utilizador
        empresa = self.request.user.funcionario.empresa
        return Lote.objects.filter(produto__empresa=empresa)

# --- CRUD COMPLETO PARA LOTES ---

class LoteListView(LoteBaseView, ListView):
    template_name = "produtos/lote_list.html"
    context_object_name = "lotes"
    paginate_by = 20

    def get_queryset(self):
        # Começa com o queryset já filtrado por empresa da LoteBaseView
        queryset = super().get_queryset().select_related('produto')
        
        search_query = self.request.GET.get('search', '').strip()
        status_filter = self.request.GET.get('status', '')
        sort_by = self.request.GET.get('sort', '-data_validade')

        if search_query:
            # CORREÇÃO: Usar 'nome_comercial'
            queryset = queryset.filter(
                Q(produto__nome_comercial__icontains=search_query) |
                Q(numero_lote__icontains=search_query)
            )

        hoje = date.today()
        if status_filter == 'vencido':
            queryset = queryset.filter(data_validade__lt=hoje)
        elif status_filter == 'vencendo':
            data_limite = hoje + timedelta(days=30)
            queryset = queryset.filter(data_validade__gte=hoje, data_validade__lte=data_limite)

        valid_sort_options = ['data_validade', '-data_validade', 'quantidade_atual', '-quantidade_atual']
        if sort_by in valid_sort_options:
            queryset = queryset.order_by(sort_by)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_sort'] = self.request.GET.get('sort', '-data_validade')
        context['add_product_form'] = ProdutoForm() 
        context['title'] = "Gestão de Lotes de Produtos"
        return context

class LoteDetailView(LoteBaseView, DetailView):
    template_name = "produtos/lote_detail.html"
    context_object_name = "lote"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalhes do Lote: {self.object.numero_lote}"
        return context

class LoteCreateView(LoteBaseView, CreateView):
    form_class = LoteForm
    template_name = "produtos/lote_form.html"
    success_url = reverse_lazy("produtos:lote_list")

    def form_valid(self, form):
        lote = form.save(commit=False)
        lote.quantidade_atual = form.cleaned_data['quantidade_inicial']
        lote.save()
        messages.success(self.request, f"Lote {lote.numero_lote} adicionado com sucesso!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Adicionar Novo Lote"
        return context

class LoteUpdateView(LoteBaseView, UpdateView):
    form_class = LoteForm
    template_name = "produtos/lote_form.html"
    
    def get_success_url(self):
        return reverse_lazy('produtos:lote_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Lote atualizado com sucesso!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Editar Lote: {self.object.numero_lote}"
        return context

class LoteDeleteView(LoteBaseView, DeleteView):
    template_name = "produtos/lote_confirm_delete.html"
    success_url = reverse_lazy("produtos:lote_list")
    
    def form_valid(self, form):
        messages.success(self.request, f"Lote '{self.object.numero_lote}' eliminado com sucesso!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Eliminar Lote"
        return context
    
# ===============================
# AJAX VIEWS
# ===============================

class BuscarProdutoAjaxView(LoginRequiredMixin, View):
    def get(self, request):
        termo = request.GET.get('q', '').strip()
        
        if len(termo) < 2:
            return JsonResponse({'produtos': []})
        
        produtos = Produto.objects.filter(
            Q(nome_produto__icontains=termo) |
            Q(codigo_barras__icontains=termo) |
            Q(codigo_interno__icontains=termo),
            ativo=True
        ).select_related('categoria', 'fabricante')[:10]
        
        produtos_data = []
        for produto in produtos:
            produtos_data.append({
                'id': produto.id,
                'nome_produto': produto.nome_produto,
                'codigo_barras': produto.codigo_barras,
                'preco_venda': float(produto.preco_venda),
                'fabricante': produto.fabricante.nome if produto.fabricante else '',
                'categoria': produto.categoria.nome if produto.categoria else '',
                'estoque_atual': produto.estoque_atual,
            })
        
        return JsonResponse({'produtos': produtos_data})


class VerificarCodigoView(LoginRequiredMixin, View):
    def get(self, request):
        codigo = request.GET.get('codigo', '').strip()
        existe = Produto.objects.filter(codigo_barras=codigo).exists()
        return JsonResponse({'existe': existe})

# ===============================
# RELATÓRIOS E CONSULTAS
# ===============================

class LotesVencimentoView(LoginRequiredMixin, ListView):
    model = Lote
    template_name = "produtos/lotes_vencimento.html"
    context_object_name = "lotes"

    def get_queryset(self):
        hoje = timezone.now().date()
        dias = int(self.request.GET.get('dias', 30))
        data_limite = hoje + timedelta(days=dias)
        
        return Lote.objects.filter(
            validade__gte=hoje,
            validade__lte=data_limite,
            quantidade__gt=0
        ).select_related('produto').order_by('validade')

class ProdutosEstoqueBaixoView(LoginRequiredMixin, ListView):
    model = Produto
    template_name = "produtos/estoque_baixo.html"
    context_object_name = "produtos"

    def get_queryset(self):
        produtos = Produto.objects.filter(ativo=True, estoque_minimo__gt=0)
        produtos_baixo = []
        
        for produto in produtos:
            if produto.estoque_baixo:
                produtos_baixo.append(produto)
        
        return produtos_baixo


# ===============================
# AÇÕES ESPECIAIS
# ===============================

class AtivarProdutoView(LoginRequiredMixin, View):
    def post(self, request, pk):
        produto = get_object_or_404(Produto, pk=pk)
        produto.ativo = True
        produto.save()
        return redirect("produtos:produto_detail", pk=pk)

class DesativarProdutoView(LoginRequiredMixin, View):
    def post(self, request, pk):
        produto = get_object_or_404(Produto, pk=pk)
        produto.ativo = False
        produto.save()
        return redirect("produtos:produto_detail", pk=pk)


@login_required
@require_http_methods(["GET"])
def listar_categorias_api(request):
    """
    API para listar categorias
    """
    try:
        if hasattr(request.user, 'funcionario') and request.user.funcionario:
            empresa = request.user.funcionario.empresa
        else:
            return JsonResponse({
                'success': False,
                'message': 'Empresa não encontrada'
            }, status=400)
        
        from .models import Categoria
        categorias = Categoria.objects.filter(empresa=empresa, ativo=True)
        
        categorias_data = []
        for categoria in categorias:
            categorias_data.append({
                'id': categoria.id,
                'nome': categoria.nome,
                'descricao': getattr(categoria, 'descricao', ''),
            })
        
        return JsonResponse({
            'success': True,
            'categorias': categorias_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }, status=500)


import os
import pandas as pd
import unicodedata
import logging
from django.views.generic.edit import FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.db.models import Q
from django.contrib import messages
from django.utils.encoding import force_str
from produtos.forms import ImportarProdutosForm
from produtos.models import Produto, Categoria, Fornecedor

logger = logging.getLogger(__name__)

class ImportarProdutosView(LoginRequiredMixin, FormView):
    template_name = 'produtos/importar_produtos.html'
    form_class = ImportarProdutosForm
    success_url = reverse_lazy('produtos:produto_list')
    acao_requerida = 'editar_produtos'

    def get_empresa(self):
        """Obtém a empresa associada ao utilizador autenticado."""
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario.empresa:
            return user.funcionario.empresa
        return None

    def sanitize_text(self, text):
        """Remove caracteres inválidos e normaliza encoding."""
        if not text:
            return ''
        try:
            text = str(text)
            text = unicodedata.normalize('NFKC', text)
            text = text.encode('utf-8', 'ignore').decode('utf-8', 'ignore')
            return text.strip()
        except Exception as e:
            logger.warning(f"Erro ao limpar texto: {text} | {str(e)}")
            return ''

    def obter_categoria(self, nome, empresa):
        """Obtém ou cria uma categoria sanitizada."""
        nome_limpo = self.sanitize_text(nome)
        if not nome_limpo:
            return None
        categoria, _ = Categoria.objects.get_or_create(
            nome__iexact=nome_limpo,
            defaults={'nome': nome_limpo, 'empresa': empresa},
            empresa=empresa
        )
        return categoria

    def obter_fornecedor(self, nome, empresa):
        """Obtém o fornecedor sanitizado ou None se não existir."""
        nome_limpo = self.sanitize_text(nome)
        if not nome_limpo:
            return None
        try:
            fornecedor = Fornecedor.objects.filter(
                Q(nome_fantasia__iexact=nome_limpo) | Q(razao_social__iexact=nome_limpo),
                empresa=empresa
            ).first()
            return fornecedor
        except Exception as e:
            logger.warning(f"Erro ao buscar fornecedor '{nome_limpo}': {str(e)}")
            return None

    def processar_arquivo(self, arquivo, empresa):
        """Processa o CSV com tolerância a erros de encoding."""
        try:
            df = pd.read_csv(arquivo, encoding='utf-8', errors='replace')
        except UnicodeDecodeError:
            df = pd.read_csv(arquivo, encoding='latin1', errors='replace')

        registros_criados, registros_atualizados = 0, 0

        for _, row in df.iterrows():
            try:
                nome = self.sanitize_text(row.get('nome'))
                preco = row.get('preco', 0)
                categoria_nome = self.sanitize_text(row.get('categoria'))
                fornecedor_nome = self.sanitize_text(row.get('fornecedor'))

                if not nome:
                    continue

                categoria = self.obter_categoria(categoria_nome, empresa)
                fornecedor = self.obter_fornecedor(fornecedor_nome, empresa)

                produto, criado = Produto.objects.update_or_create(
                    nome__iexact=nome,
                    empresa=empresa,
                    defaults={
                        'nome': nome,
                        'preco': preco,
                        'categoria': categoria,
                        'fornecedor': fornecedor,
                    },
                )

                if criado:
                    registros_criados += 1
                else:
                    registros_atualizados += 1

            except Exception as e:
                logger.error(f"Erro ao processar linha {row.to_dict()}: {str(e)}")

        return registros_criados, registros_atualizados

    def form_valid(self, form):
        empresa = self.get_empresa()
        if not empresa:
            messages.error(self.request, "Empresa não associada ao utilizador.")
            return self.form_invalid(form)

        arquivo = form.cleaned_data.get('arquivo')
        if not arquivo:
            messages.error(self.request, "Nenhum arquivo CSV foi enviado.")
            return self.form_invalid(form)

        try:
            registros_criados, registros_atualizados = self.processar_arquivo(arquivo, empresa)
            messages.success(
                self.request,
                f"Importação concluída com sucesso. {registros_criados} produtos criados, {registros_atualizados} atualizados."
            )
        except Exception as e:
            logger.exception(f"Erro geral ao importar produtos: {str(e)}")
            messages.error(self.request, f"Erro ao importar produtos: {str(e)}")

        return super().form_valid(form)

   
    